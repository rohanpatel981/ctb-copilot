"""Excel export — turn a QueryResult dict into a 3-sheet .xlsx workbook.

Used by the Streamlit UI's "Export to Excel" button. Pure data-shaping: no
LLM, no config, no DB. Takes a result dict (same shape as run_query returns)
and produces bytes suitable for st.download_button or any HTTP response.

The output workbook has:
  - Sheet "Answer"        — question, confidence, explanation, YoY/ratio breakdowns
  - Sheet "SQL"           — the exact SQL Claude generated, monospaced
  - Sheet "Source Rows"   — every row that came back from DuckDB
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


_HEADER_FILL = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
_LABEL_WIDTH = 20
_VALUE_WIDTH = 60
_MAX_AUTOFIT_WIDTH = 50


def _bold(size: int = 11) -> Font:
    return Font(bold=True, size=size)


def _title_cell(ws: Worksheet, text: str) -> None:
    ws["A1"] = text
    ws["A1"].font = _bold(14)


def _kv(ws: Worksheet, row: int, label: str, value: Any) -> None:
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = _bold()
    label_cell.alignment = Alignment(vertical="top")
    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.alignment = Alignment(wrap_text=True, vertical="top")


def _populate_answer_sheet(ws: Worksheet, result: dict) -> None:
    _title_cell(ws, "ctb-copilot — answer")

    _kv(ws, 3, "Question", result.get("question", ""))
    _kv(ws, 4, "Confidence", result.get("confidence", ""))
    _kv(ws, 5, "Post-processing", result.get("post_process", "none"))
    _kv(ws, 6, "Generated", datetime.now().isoformat(timespec="seconds"))
    _kv(ws, 8, "Explanation", result.get("explanation", ""))

    row = 10
    if result.get("yoy_changes"):
        ws.cell(row=row, column=1, value="YoY Changes").font = _bold()
        row += 1
        for col_idx, header in enumerate(
            ["Metric", "From period", "From value", "To period", "To value", "% change"], start=1
        ):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = _bold()
            cell.fill = _HEADER_FILL
        row += 1
        for ch in result["yoy_changes"]:
            ws.cell(row=row, column=1, value=ch.get("metric"))
            ws.cell(row=row, column=2, value=ch.get("from_period"))
            ws.cell(row=row, column=3, value=ch.get("from_value"))
            ws.cell(row=row, column=4, value=ch.get("to_period"))
            ws.cell(row=row, column=5, value=ch.get("to_value"))
            ws.cell(row=row, column=6, value=ch.get("pct_change"))
            row += 1
        row += 1

    if result.get("ratios"):
        ws.cell(row=row, column=1, value="Ratios").font = _bold()
        row += 1
        for col_idx, header in enumerate(["Period", "Numerator", "Denominator", "Value"], start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = _bold()
            cell.fill = _HEADER_FILL
        row += 1
        for r in result["ratios"]:
            ws.cell(row=row, column=1, value=r.get("period"))
            ws.cell(row=row, column=2, value=r.get("numerator"))
            ws.cell(row=row, column=3, value=r.get("denominator"))
            ws.cell(row=row, column=4, value=r.get("value"))
            row += 1

    ws.column_dimensions["A"].width = _LABEL_WIDTH
    ws.column_dimensions["B"].width = _VALUE_WIDTH
    for c in ("C", "D", "E", "F"):
        ws.column_dimensions[c].width = 18


def _populate_sql_sheet(ws: Worksheet, result: dict) -> None:
    _title_cell(ws, "ctb-copilot — generated SQL")

    sql_cell = ws.cell(row=3, column=1, value=result.get("sql", ""))
    sql_cell.alignment = Alignment(wrap_text=True, vertical="top")
    sql_cell.font = Font(name="Courier New", size=10)

    ws.column_dimensions["A"].width = 120
    sql_text = result.get("sql", "") or ""
    line_count = max(sql_text.count("\n") + 1, 6)
    ws.row_dimensions[3].height = max(20 * line_count, 80)

    pp = result.get("post_process", "none")
    if pp != "none":
        note = ws.cell(row=5, column=1, value=f"Post-processing applied: {pp}")
        note.font = _bold()


def _populate_rows_sheet(ws: Worksheet, result: dict) -> None:
    _title_cell(ws, "ctb-copilot — source rows")

    columns = result.get("columns") or []
    rows = result.get("rows") or []

    if not columns or not rows:
        ws["A3"] = "(no rows returned)"
        return

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = _bold()
        cell.fill = _HEADER_FILL

    for row_idx, row_dict in enumerate(rows, start=4):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(col_name))

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            [len(str(col_name))]
            + [len(str(r.get(col_name, ""))) for r in rows[:100]]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, _MAX_AUTOFIT_WIDTH)

    ws.freeze_panes = "A4"


def query_result_to_xlsx(result: dict) -> bytes:
    """Build a 3-sheet .xlsx (Answer / SQL / Source Rows) from a QueryResult dict.

    Returns raw bytes — feed straight to st.download_button or write to disk.
    """
    wb = Workbook()
    answer_sheet = wb.active
    answer_sheet.title = "Answer"
    _populate_answer_sheet(answer_sheet, result)

    sql_sheet = wb.create_sheet("SQL")
    _populate_sql_sheet(sql_sheet, result)

    rows_sheet = wb.create_sheet("Source Rows")
    _populate_rows_sheet(rows_sheet, result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(question: str | None = None) -> str:
    """Suggested filename for downloads."""
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"ctb-copilot-{ts}.xlsx"
