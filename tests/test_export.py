"""Tests for the Excel export. No LLM, no DB, no config."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from ctb_copilot.export import export_filename, query_result_to_xlsx


def _result(**overrides: Any) -> dict:
    base = {
        "question": "What's the total of assets for FY 2024-25?",
        "sql": "SELECT period, fs_category, SUM(amount_consolidated) AS total\nFROM ctb_data\nWHERE fs_category='Assets' AND period='FY 2024-25'\nGROUP BY period, fs_category",
        "explanation": "Sums amount_consolidated across Asset rows for FY 2024-25.",
        "confidence": "high",
        "post_process": "none",
        "columns": ["period", "fs_category", "total"],
        "rows": [{"period": "FY 2024-25", "fs_category": "Assets", "total": 272564694.79}],
        "yoy_changes": [],
        "ratios": [],
    }
    base.update(overrides)
    return base


def _open(bytes_: bytes):
    return load_workbook(io.BytesIO(bytes_), data_only=False)


def test_returns_valid_workbook_with_three_sheets() -> None:
    wb = _open(query_result_to_xlsx(_result()))
    assert wb.sheetnames == ["Answer", "SQL", "Source Rows"]


def test_answer_sheet_contains_question_and_confidence() -> None:
    wb = _open(query_result_to_xlsx(_result()))
    ws = wb["Answer"]
    assert ws["A1"].value == "ctb-copilot — answer"
    assert ws["A3"].value == "Question"
    assert "total of assets" in ws["B3"].value
    assert ws["A4"].value == "Confidence"
    assert ws["B4"].value == "high"


def test_sql_sheet_preserves_sql_text() -> None:
    wb = _open(query_result_to_xlsx(_result()))
    ws = wb["SQL"]
    assert "SELECT period" in ws["A3"].value
    assert "fs_category='Assets'" in ws["A3"].value


def test_source_rows_sheet_has_headers_and_data() -> None:
    wb = _open(query_result_to_xlsx(_result()))
    ws = wb["Source Rows"]
    assert ws["A3"].value == "period"
    assert ws["B3"].value == "fs_category"
    assert ws["C3"].value == "total"
    assert ws["A4"].value == "FY 2024-25"
    assert ws["B4"].value == "Assets"
    assert abs(ws["C4"].value - 272564694.79) < 0.01


def test_source_rows_sheet_handles_empty_result() -> None:
    wb = _open(query_result_to_xlsx(_result(rows=[], columns=[])))
    ws = wb["Source Rows"]
    assert ws["A3"].value == "(no rows returned)"


def test_yoy_changes_render_when_present() -> None:
    yoy = [
        {
            "metric": "total_assets",
            "from_period": "FY 2024-25",
            "from_value": 100.0,
            "to_period": "FY 2025-26",
            "to_value": 130.0,
            "pct_change": 30.0,
        }
    ]
    wb = _open(query_result_to_xlsx(_result(post_process="yoy_pct", yoy_changes=yoy)))
    ws = wb["Answer"]
    found = False
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell == "YoY Changes":
                found = True
    assert found, "Answer sheet should include a 'YoY Changes' section"


def test_ratios_render_when_present() -> None:
    ratios = [{"period": "FY 2024-25", "numerator": "current_assets", "denominator": "current_liabilities", "value": 1.5}]
    wb = _open(query_result_to_xlsx(_result(post_process="ratio", ratios=ratios)))
    ws = wb["Answer"]
    found = any(cell == "Ratios" for row in ws.iter_rows(values_only=True) for cell in row)
    assert found, "Answer sheet should include a 'Ratios' section"


def test_export_filename_has_xlsx_extension_and_timestamp() -> None:
    name = export_filename("any question")
    assert name.endswith(".xlsx")
    assert name.startswith("ctb-copilot-")
    assert len(name) > len("ctb-copilot-.xlsx")


def test_round_trip_preserves_numeric_types() -> None:
    rows = [{"period": "FY 2024-25", "value": 12345.67}, {"period": "FY 2025-26", "value": 23456.78}]
    wb = _open(query_result_to_xlsx(_result(rows=rows, columns=["period", "value"])))
    ws = wb["Source Rows"]
    assert isinstance(ws["B4"].value, float)
    assert isinstance(ws["B5"].value, float)
